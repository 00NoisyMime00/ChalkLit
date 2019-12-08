import Levenshtein as lev
from openpyxl import load_workbook

wb=load_workbook("Physics_Data_Untagged(Large).xlsx")
ws=wb.active

ontology={}

for row in ws.iter_rows(min_row=2, max_row=2198, min_col=1, max_col=5, values_only=True):
	i=0
	entity=None
	attribute=[]
	for col in reversed(row):
		i+=1
		if i==1:
			if col==None or col=="[Request Failed]":
				break
			entity=col
			if entity not in ontology:
				ontology[entity]={"Definition": [], "Types": [], "Effects": [], "Examples": [], "Application": [], "Property": [], "Causes": [], "Reasoning": [], "Formulae": []}
		elif i==4:
			attribute=col.strip().split(",")
			j=0
			for a in attribute:
				match=(100, None)
				for k in ontology[entity].keys():
					dist=lev.distance(a.lower(),k.lower())
					if(dist<match[0]):
						match=(dist, k)
				attribute.pop(j)
				attribute.insert(j, match[1])
				j+=1
		elif i==5:
			for u in attribute:
				ontology[entity][u].append(col)
k=0
for i in ontology["Refractive index"]["Formulae"]:
	k+=1
	print("%d. %s" % (k,i))
'''
#f=open("check_entity", "w+")				
for o in ontology.keys():
	#f.write(o)
	#f.write("\n")
	print(o)
	for a in ontology[o].keys():
		print(a)
		print(ontology[o][a])
	print("")
#f.close()'''