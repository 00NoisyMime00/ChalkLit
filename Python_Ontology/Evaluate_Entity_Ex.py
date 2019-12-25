import Levenshtein as lev, codecs, pickle
from openpyxl import Workbook

### CAUTION- Don't change the below snippet to load the ontology!
with codecs.open("ontology", "r", encoding="utf-8", errors="ignore") as f: # Don't load the ontology using just open 
    ontology=pickle.load(f) # Loads the Ontology

wb=Workbook()
ws=wb.active
ws.title="Entity Extractor Evaluation"
ws.cell(row=1, column=1, value="Question")
ws.cell(row=1, column=2, value="Extracted Entity")
ws.cell(row=1, column=3, value="Actual Entity") # This is Error Analysis, so evaluate it yourself and write in the provided column

def extract_entities(question): 
	"""
		Ask questions like-
			Whats force?
			Why Air Pressure differnce causes Tornado?
		Most importantly ask as many different types of questions as possible

		To evaluate use the Confusion Matrix, read more here "https://machinelearningmastery.com/classification-accuracy-is-not-enough-more-performance-measures-you-can-use/"
		Note: The list of all the Entities in our ontology is in all_distinct.txt, use it to fill the Actual Entity column 
		
		If you find a way to improve the performance, make changes in the extract_entity(question) function which 
		does the entity extraction part, firstly in the same file not in EduQA_v2 in a new 
		Branch and name the Branch Evaluate_Entity_Ex_YourName, we'll evaluate all the models again and keep the one
		with the best performance

		Please do atleast 50 questions and evaluate all of them, won't take more than a couple of hours. Make 
		changes in the model ie. extract_entities(question) function to improve the performance, we'll discuss the 
		performance today night. 
	"""
	question.replace("?", "")
	order=set()
	with open("entity_ex_help.txt", "r") as help:
		words=question.split()
		for lines in help:
			for w in words:	
				if lev.distance(lines.strip().lower(), w.lower())<2:
					order.add(lines.strip())		
	print(order)	
	entity=None
	max=(0, 10)				
	with open("all_distinct.txt", "r") as all:
		for lines in all:
			words=lines.strip().split()
			match=0
			for o in order:
				if o in words:
					match+=1
			if match>0:
				if match==max[0] and len(words)<max[1] or match>max[0]:
					max=(match, len(words))
					entity=lines.strip()
	print(entity)
	return entity

i=2
while True:
	question=input("Enter your question: ").lower() 
	if question=="exit":
		break
	entity=extract_entities(question)             
	ws.cell(row=i, column=1, value=question)
	ws.cell(row=i, column=2, value=entity)
	i+=1

wb.save("Entity_Ex_Eval.xlsx")