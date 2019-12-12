"""
	onto_pop.py populates the ontology and dumps it into a file named ontology
"""

import Levenshtein as lev, pickle # Used for spelling error correction of question type Eg: Definion -> Definition
from openpyxl import load_workbook

wb=load_workbook("Physics_Data_Untagged(Large).xlsx")
ws=wb.active

ontology={} # This is our ontology which is later dumped
			# Key: entity
			# Value: Another Dictionary
				# Key: Question Type/Attribute
				# Value: Corresponding Paragraphs

for row in ws.iter_rows(min_row=2, max_row=2155, min_col=1, max_col=6, values_only=True):
	i=0
	entity=[] # Since a paragraph can correspond to more than 1 entity
	attribute=[] # Or Question Type
	for col in reversed(row):
		i+=1
		if i==1:
			if col==None:
				break
			entity=col.strip().split(",") # Stores all entities corresponding to the paragraph
			for e in entity:
				if e not in ontology: # A value is created corresponding to the key which is the entity corresponding to the paragraph
					ontology[e]={"Definition": [], "Types": [], "Effects": [], "Examples": [], "Application": [], "Property": [], "Causes": [], "Reasoning": [], "Formulae": []}
		elif i==5:
			attribute=col.strip().split(",")
			j=0
			for a in attribute:
				match=(100, None)
				for k in ontology[entity[0]].keys():
					dist=lev.distance(a.lower(),k.lower()) # Spelling errors are handled here see 'https://www.datacamp.com/community/tutorials/fuzzy-string-python' for more info
					if(dist<match[0]):
						match=(dist, k)
				attribute.pop(j)
				attribute.insert(j, match[1])
				j+=1
		elif i==6:
			for u in attribute:
				for e in entity:
					ontology[e][u].append(col) # Each entity corresponding to the paragraph is populated with the paragraph in respective Question Type

f=open("ontology", "wb")
pickle.dump(ontology, f) # Serializes ontology
f.close()

# Below is just to check is ontology is being populated "Uniquely ie. no Duplication"

'''
f=open("check_entity.txt", "w+")				
for o in ontology.keys():
	f.write(o)
	f.write("\n")
f.close()

lines_seen = set() # holds lines already seen
dup=[]
outfile = open("all_distinct.txt", "w")
for line in open("check_entity.txt", "r"):
    if line not in lines_seen: # not a duplicate
        outfile.write(line)
        lines_seen.add(line)
    else:
    	dup.append(line)
outfile.write("\n")
for d in dup:
	outfile.write(d)
outfile.close()

# Well, below code is just for re-conformation

with open("all_distinct.txt", "r") as file1:
    with open("check_entity.txt", "r") as file2:
        diff = set(file1).symmetric_difference(file2)
diff.discard("\n")

with open("some_output_file.txt", "w") as file_out:
    for line in diff:
        file_out.write(line)
'''