'''
	ExcelPrePro.py preprocesses the Excel Sheet- since the way of tagging adopted by everyone was different
	Note: Now onwards, whenever we have to tag we will first adopt a standard which should be followed while tagging,
		it took me 2 days for debugging just to realize in the end that the error is caused due to different tagging
		strategy adopted by everyone- 
			Eg: If two entities correspond to the same paragraph, someone has separated them
			using ',' whereas others have used ';' or '\n'
'''

from openpyxl import load_workbook

wb=load_workbook("Physics_Data_Untagged(Large).xlsx")
ws=wb.active

for row in ws.iter_rows(min_row=2, max_row=2155, min_col=6, max_col=6): # Only column 6 has the tagged entities, thus only it should be preprocessed to make it uniform 
	for col in row:
		value=col.value
		if value!=None:
			value=value.replace("_", " ")
			value=value.replace(";",",")
			value=value.replace("/",",")
			value=value.replace(", ",",")
			value=value.replace("\n",",")
			value=value.lstrip(" ")
			value=value.rstrip(" ")
			value=value.lower()
			col.value=value

wb.save("Physics_Data_Untagged(Large).xlsx")